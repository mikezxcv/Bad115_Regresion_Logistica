#Librerías para el manejo de datos y generación de gráficos
# ============================================================================== 
import pandas as pd
import numpy as np
import openpyxl
import csv
import os

from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import statsmodels.api as sm

import matplotlib.pyplot as plt
from matplotlib import style
import seaborn as sns
import matplotlib.pyplot as plt
plt.rcParams['image.cmap'] = "bwr"
plt.rcParams['savefig.bbox'] = "tight"
style.use('ggplot') or plt.style.use('ggplot')

# ==============================================================================

#Procesamiento de los datos
# ==============================================================================
#Lectura de los datos
dataframe = pd.read_csv(r"water_potability.csv")

fill_na = 0

if fill_na == 1:
    dataframe['ph']= dataframe['ph'].fillna(dataframe['ph'].mean())
    dataframe['Hardness']= dataframe['Hardness'].fillna(dataframe['Hardness'].mean())
    dataframe['Chloramines']= dataframe['Chloramines'].fillna(dataframe['Chloramines'].mean())
    dataframe['Sulfate']= dataframe['Sulfate'].fillna(dataframe['Sulfate'].mean())
    dataframe['Conductivity']= dataframe['Conductivity'].fillna(dataframe['Conductivity'].mean())
    dataframe['Organic_carbon']= dataframe['Organic_carbon'].fillna(dataframe['Organic_carbon'].mean())
    dataframe['Trihalomethanes']= dataframe['Trihalomethanes'].fillna(dataframe['Trihalomethanes'].mean())
    dataframe['Turbidity']= dataframe['Turbidity'].fillna(dataframe['Turbidity'].mean())
   
dataframe = dataframe.dropna()

x = dataframe.drop(columns = 'Potability')
y = dataframe['Potability']

#División del conjunto de datos para entrenamiento y testeo
x_train, x_test, y_train, y_test = train_test_split(
                                        x,
                                        y.values.reshape(-1,1),
                                        train_size   = 0.8,
                                        random_state = 1,
                                        shuffle      = True
                                    )

#Creación del modelo
x_train = sm.add_constant(x_train, prepend=True)
modelo = sm.Logit(endog=y_train, exog=x_train,)
modelo = modelo.fit()

#Almacenamiento del resumen de los resultados del modelo
modelo_resumen = modelo.summary().as_csv()
write_path = './modelo_resumen.csv'
with open(write_path, 'w') as f:
    f.write(modelo_resumen)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Modelo_Resumen"
with open('modelo_resumen.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)

#Cálculo de las predicciones sobre datos de testeo
x_test = sm.add_constant(x_test, prepend=True)
predicciones = modelo.predict(exog = x_test)
clasificacion = np.where(predicciones<0.5, 0, 1)

#Cálculo de la precisión del modelo
accuracy = accuracy_score(
            y_true    = y_test,
            y_pred    = clasificacion,
            normalize = True
           )

ws1 = wb.create_sheet(title="Test_Precision")
ws1.append(["El test de precisión es: ",100*accuracy,"%"])

#Generación de matriz de confusión de las predicciones
confusion_matrix = pd.crosstab(
    y_test.ravel(),
    clasificacion,
    rownames=['Real'],
    colnames=['Predicción']
)

#Almacenamiento de matriz de confusión
confusion_matrix.to_csv("./matriz_confusion.csv")

ws2 = wb.create_sheet(title="Matriz_Confusion")
ws2.append([])
with open('matriz_confusion.csv', 'r') as f:
    for row in csv.reader(f):
        ws2.append(row)
ws2['B1'] = "Predicción"
ws2['A7'] = "VN"
ws2['B7'] = "Verdadero Negativo"
ws2['A8'] = "FN"
ws2['B8'] = "Falso Negativo"
ws2['A9'] = "FP"
ws2['B9'] = "Falso Positivo"
ws2['A10'] = "VP"
ws2['B10'] = "Verdadero Positivo"

#Generación png de matriz de confusión
plt.clf()
plt.imshow(confusion_matrix, interpolation='nearest', cmap=plt.cm.Wistia)
classNames = ['0','1']
plt.title('Potabilidad del agua - Matriz de Confusión - Datos de testeo')
plt.ylabel('Reales')
plt.xlabel('Predicción')
tick_marks = np.arange(len(classNames))
plt.xticks(tick_marks, classNames)
plt.yticks(tick_marks, classNames)
s = [['VN','FN'], ['FP', 'VP']]
for j in range(2):
    for i in range(2):
        plt.text(i,j, str(s[i][j])+" = "+str(confusion_matrix[i][j]))

plt.savefig("matriz_confusion.png")

img = openpyxl.drawing.image.Image('matriz_confusion.png')
img.anchor = 'F1'
ws2.add_image(img)

wb.save('modelo_resumen.xlsx')
print("\nArchivo generado exitosamente.")

#Eliminación de archivos utilizados
if os.path.exists("matriz_confusion.csv"):
    os.remove("matriz_confusion.csv")
if os.path.exists("modelo_resumen.csv"):
    os.remove("modelo_resumen.csv")
if os.path.exists("matriz_confusion.png"):
    os.remove("matriz_confusion.png")

# ==============================================================================